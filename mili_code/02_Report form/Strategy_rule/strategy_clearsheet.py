# -*- coding: utf-8 -*-
"""
Created on Wed May  3 10:57:10 2017

@author: Administrator
"""

import win32com.client

filename= r'F:\celueji\StrategyReport_template.xlsx'

def vfh.Worksheets(sheetname):
 
    exc = win32com.client.Dispatch('Excel.Application')
    vfh = exc.Workbooks.Open(filename)
    sht = vfh.Worksheets(sheetname)
    for i in range(1,60):
        sht.Rows(i).Delete() 

def vfh.Getsheets(sheetname):
    for sheetname in sheetnames:
        vfh.Worksheets(sheetname)
        
if __name__=='__main__':     
    
    sheetnames = [u'Strategy﻿策略漏斗',u'RELOAN_BQS策略集',u'RELOAN_TD策略集',u'RELOAN_SIMPLE_BQS策略集',u'LOAN_BQS策略集',u'LOAN_TD策略集',u'LOAN_CX策略集']    
    vdh.Getsheets(sheetnames)
    
    
    
    
import openpyxl  
dest_filename=r'C:\Users\Administrator\Desktop\mili_mtd 2.xlsx'

def clear_sheet(Sheet2):
    '''
    Clear all cells starting from first row. we keep our header and
    replaces other rows content by None
    '''
    wb=openpyxl.load_workbook(dest_filename, read_only=False, keep_vba=True)
    ws=wb.active
    dst_ws=dst_wb.get_sheet_by_name("Sheet2")
    #We kee the header located in row 2, and set the rest to None
    if dst_ws.cell('A1').value is not None: #check if already cleared content
        for row in dst_ws.iter_rows(row_offset=2):
            for cell in row:
                cell.value=None
    wb.save(dest_filename)
    
def copy_data(searchMilestone):
    '''
    gets data from the sequence file
    '''
    dst_wb=openpyxl.load_workbook(dest_filename, read_only=False, keep_vba=True)
    dst_ws=dst_wb.get_sheet_by_name(sequence_sheet)
    data =[]
    print dst_ws.max_row #here should return 3, but it seems it is counting all the cells we set to None
    for i in xrange(sequenceWs.nrows):
        #this reading part is done using xlrd module
        milestone=sequenceWs.cell(i,1).value
        execution=sequenceWs.cell(i,2).value
        system=sequenceWs.cell(i,16).value
        if searchMilestone in milestone and 'WC' in execution:
            #copy some data
            line=[data, data, data]
            dst_ws.append(line)
    dst_wb.save(dest_filename)
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
import asposecellscloud
from asposecellscloud.CellsApi import CellsApi
from asposecellscloud.CellsApi import ApiException

import asposestoragecloud
from asposestoragecloud.StorageApi import StorageApi

apiKey = "XXXXX" #sepcify App Key
appSid = "XXXXX" #sepcify App SID
apiServer = "http://api.aspose.com/v1.1"
data_folder = "../../data/"

#Instantiate Aspose Storage API SDK
storage_apiClient = asposestoragecloud.ApiClient.ApiClient(apiKey, appSid, True)
storageApi = StorageApi(storage_apiClient)
#Instantiate Aspose Cells API SDK
api_client = asposecellscloud.ApiClient.ApiClient(apiKey, appSid, True)
cellsApi = CellsApi(api_client);

#set input file name
filename = "Sample_Test_Book.xls"
sheetName = "Sheet1"
startRow = 1
startColumn = 1
endRow = 2
endColumn = 2

#upload file to aspose cloud storage
storageApi.PutCreate(Path=filename, file=data_folder + filename)

try:
    #invoke Aspose.Cells Cloud SDK API to clear contents and styles of selected cells in a worksheet
    response = cellsApi.PostClearContents(name=filename, sheetName=sheetName, startRow=startRow, startColumn=startColumn, endRow=endRow, endColumn=endColumn)

    if response.Status == "OK":
        #download updated Workbook from storage server
        response = storageApi.GetDownload(Path=filename)
        outfilename = "c:/temp/" + filename
        with open(outfilename, 'wb') as f:
                    for chunk in response.InputStream:
                        f.write(chunk)


except ApiException as ex:
            print "ApiException:"
            print "Code:" + str(ex.code)
            print "Message:" + ex.message    