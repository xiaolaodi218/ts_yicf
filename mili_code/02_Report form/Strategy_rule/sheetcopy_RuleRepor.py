# -*- coding: utf-8 -*-
"""
Created on Sat Mar 25 16:08:16 2017

@author: lenovo
"""
import time
start =time.clock()

from openpyxl import load_workbook

def replace_xls(sheetname):
    filename = 'F:\\celueji\\sas_csv\RuleReport.xlsx'
    filename2 = 'F:\\celueji\\RuleReport_template.xlsx'

    wb = load_workbook(filename)
    wb2 = load_workbook(filename2)
    ws = wb[sheetname]
    ws2 = wb2[sheetname]

    for i,row in enumerate(ws.iter_rows()):
        for j,cell in enumerate(row):
            ws2.cell(row=i+1, column=j+1, value=cell.value)

    wb2.save(filename2)

#sheetnames = ['新客户订单氪信规则']
#sheetnames = ['复贷客户订单白骑士规则','复贷2客户订单白骑士规则','新客户订单白骑士规则']
sheetnames = [
'复贷客户订单白骑士规则',
'复贷2客户订单白骑士规则',
'新客户订单白骑士规则',
'复贷客户订单复贷准入',
'复贷客户订单复贷基本规则',
'复贷客户订单FSYYS规则',
'复贷客户订单FSDS规则',
'复贷客户订单关联规则',
'复贷客户订单同盾规则',
'复贷2客户订单复贷准入',
'复贷2客户订单复贷基本规则',
'复贷2客户订单FSYYS规则',
'复贷2客户订单FSDS规则',
'复贷2客户订单关联规则',
'新客户订单基本规则',
'新客户订单FSDS规则',
'新客户订单关联规则',
'新客户订单通过规则',
'新客户订单氪信规则',
'新客户订单同盾规则',]

for sheetname in sheetnames:
    replace_xls(sheetname)

#==============================================================================
#贷款事件FSYYS冠军规则
def replace_xls(sheetname):
    filename = 'F:\\celueji\\sas_csv\RuleReport_a.xlsx'
    filename2 = 'F:\\celueji\\RuleReport_template.xlsx'

    wb = load_workbook(filename)
    wb2 = load_workbook(filename2)
    ws = wb[sheetname]
    ws2 = wb2[sheetname]

    for i,row in enumerate(ws.iter_rows()):
        for j,cell in enumerate(row):
            ws2.cell(row=i+1, column=j+1, value=cell.value)

    wb2.save(filename2)

sheetnames = ['新客户订单FSYYS规则']

for sheetname in sheetnames:
    replace_xls(sheetname)
    
#==============================================================================
#贷款事件FSYYS挑战者规则
def replace_xls(sheetname):
    filename = 'F:\\celueji\\sas_csv\RuleReport_b.xlsx'
    filename2 = 'F:\\celueji\\RuleReport_template.xlsx'

    wb = load_workbook(filename)
    wb2 = load_workbook(filename2)
    ws = wb[sheetname]
    ws2 = wb2[sheetname]

    for i,row in enumerate(ws.iter_rows()):
        for j,cell in enumerate(row):
            ws2.cell(row=i+1, column=j+1, value=cell.value)

    wb2.save(filename2)

sheetnames = ['新客户订单FSYYS_B规则']

for sheetname in sheetnames:
    replace_xls(sheetname)
    
#==============================================================================
#决策事件 基本规则 FSYYS规则；冠军
def replace_xls(sheetname):
    filename = 'F:\\celueji\\sas_csv\RuleReport1_a.xlsx'
    filename2 = 'F:\\celueji\\RuleReport_template.xlsx'

    wb = load_workbook(filename)
    wb2 = load_workbook(filename2)
    ws = wb[sheetname]
    ws2 = wb2[sheetname]

    for i,row in enumerate(ws.iter_rows()):
        for j,cell in enumerate(row):
            ws2.cell(row=i+1, column=j+1, value=cell.value)

    wb2.save(filename2)

sheetnames = [u'新客户订单基本规则_冠军']

for sheetname in sheetnames:
    replace_xls(sheetname)

#==============================================================================
#决策事件 基本规则_b FSYYS规则_b  挑战者
def replace_xls(sheetname):
    filename = 'F:\\celueji\\sas_csv\RuleReport1_b.xlsx'
    filename2 = 'F:\\celueji\\RuleReport_template.xlsx'

    wb = load_workbook(filename)
    wb2 = load_workbook(filename2)
    ws = wb[sheetname]
    ws2 = wb2[sheetname]

    for i,row in enumerate(ws.iter_rows()):
        for j,cell in enumerate(row):
            ws2.cell(row=i+1, column=j+1, value=cell.value)

    wb2.save(filename2)

#sheetnames = ['新客户订单基本规则_挑战者','新客户订单FSYYS_B规则_挑战者']
sheetnames = ['新客户订单基本规则_挑战者']

for sheetname in sheetnames:
    replace_xls(sheetname)
    
    
end = time.clock()
print('Running time: %s Seconds'%(end-start))
 
