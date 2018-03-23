# -*- coding: utf-8 -*-
"""
Created on Tue Apr 11 11:42:41 2017

@author: Administrator
"""
import time
start =time.clock()

from openpyxl import load_workbook

filename = r'F:\celueji\sas_csv\StrategyReport.xlsx'
filename2 = r'F:\celueji\StrategyReport_template.xlsx'

def replace_xls(sheetname):
    
    wb = load_workbook(filename)
    wb2 = load_workbook(filename2)

    ws = wb[sheetname]
    ws2 = wb2[sheetname]
    
    for i,row in enumerate(ws.iter_rows()):
        for j,cell in enumerate(row):
            ws2.cell(row=i+1, column=j+1, value=cell.value)

    wb2.save(filename2)

#sheetnames = [u'Strategy﻿策略漏斗']
sheetnames = [
u'Strategy﻿策略漏斗',
u'RELOAN_BQS策略集',
u'RELOAN_TD策略集',
u'RELOAN_SIMPLE_BQS策略集',
u'LOAN_BQS策略集',
u'LOAN_TD策略集',
u'LOAN_CX策略集']

for sheetname in sheetnames:
    replace_xls(sheetname)
    

end = time.clock()
print('Running time: %s Seconds'%(end-start))
    